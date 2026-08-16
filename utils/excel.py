import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import csv
import os
from config import EXCEL_FILE, CSV_FILE
from utils.logger import logger

def save_to_excel(movies_data, file_path=EXCEL_FILE):
    """Save movie data to Excel file without pandas dependency"""
    try:
        if not movies_data:
            logger.warning("No data to save to Excel")
            return False
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Movies 2026"
        
        # Define headers from first movie's keys
        headers = list(movies_data[0].keys())
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Add data rows
        for row_idx, movie in enumerate(movies_data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = movie.get(header, '')
                # Handle None values
                if value is None:
                    value = ''
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_idx)
            
            # Find max length in column
            max_length = len(str(header))
            for row_idx in range(2, len(movies_data) + 2):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    cell_length = len(str(cell_value))
                    if cell_length > max_length:
                        max_length = min(cell_length, 50)  # Cap at 50 characters
            
            # Set column width
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Set row height for better readability
        for row_idx in range(1, len(movies_data) + 2):
            ws.row_dimensions[row_idx].height = 20
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Save Excel file
        wb.save(file_path)
        logger.info(f"Data saved to Excel: {file_path}")
        
        # Also save as CSV
        save_to_csv(movies_data, CSV_FILE)
        
        return True
        
    except Exception as e:
        logger.error(f"Error saving to Excel: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False

def save_to_csv(movies_data, file_path=CSV_FILE):
    """Save movie data to CSV file"""
    try:
        if not movies_data:
            logger.warning("No data to save to CSV")
            return False
        
        headers = list(movies_data[0].keys())
        
        with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=headers)
            writer.writeheader()
            
            for movie in movies_data:
                # Handle None values
                row = {k: (v if v is not None else '') for k, v in movie.items()}
                writer.writerow(row)
        
        logger.info(f"Data saved to CSV: {file_path}")
        return True
        
    except Exception as e:
        logger.error(f"Error saving to CSV: {e}")
        return False

def load_from_excel(file_path):
    """Load movie data from Excel file"""
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Get headers
        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                headers.append(header)
        
        # Get data
        data = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                row_data[header] = value
            data.append(row_data)
        
        logger.info(f"Loaded data from Excel: {file_path}")
        return data
        
    except Exception as e:
        logger.error(f"Error loading from Excel: {e}")
        return None